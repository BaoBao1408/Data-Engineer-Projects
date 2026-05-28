"""
Excel Extractor – extract structured data from .xlsx / .xls files.
Handles multi-sheet workbooks, merged cells, header detection, and type inference.
Typical sources: financial models, audit schedules, client data exports.
"""
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass
class SheetData:
    sheet_name: str
    dataframe: pd.DataFrame
    row_count: int
    col_count: int
    headers: list[str]
    has_merged_cells: bool = False
    named_ranges: list[str] = field(default_factory=list)

    @property
    def is_empty(self) -> bool:
        return self.dataframe.empty


@dataclass
class ExcelDocument:
    file_path: str
    file_name: str
    file_size_bytes: int
    sheet_names: list[str]
    sheets: dict[str, SheetData]
    named_ranges: dict[str, Any]
    extraction_status: str = "success"
    error_message: Optional[str] = None


class ExcelExtractor:
    """
    Extract structured data from Excel workbooks.

    Features:
    - Multi-sheet extraction
    - Merged cell unmerging
    - Automatic header row detection
    - Data type inference (dates, numerics, categories)
    - Named range extraction
    """

    def __init__(
        self,
        sheets: Optional[list[str]] = None,
        skip_empty_sheets: bool = True,
        header_row_detection: bool = True,
        dtype_inference: bool = True,
    ):
        self.sheets = sheets          # None = all sheets
        self.skip_empty_sheets = skip_empty_sheets
        self.header_row_detection = header_row_detection
        self.dtype_inference = dtype_inference

    def extract(self, file_path: str | Path) -> ExcelDocument:
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        file_size = file_path.stat().st_size
        logger.info(f"Extracting Excel: {file_path.name} ({file_size:,} bytes)")

        try:
            wb = load_workbook(file_path, read_only=False, data_only=True)
            target_sheets = self.sheets or wb.sheetnames
            named_ranges = self._extract_named_ranges(wb)

            sheets_data: dict[str, SheetData] = {}

            for sheet_name in target_sheets:
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"Sheet '{sheet_name}' not found, skipping")
                    continue

                ws = wb[sheet_name]
                sheet_data = self._extract_sheet(ws, sheet_name)

                if self.skip_empty_sheets and sheet_data.is_empty:
                    logger.debug(f"Skipping empty sheet: {sheet_name}")
                    continue

                sheets_data[sheet_name] = sheet_data
                logger.info(
                    f"Sheet '{sheet_name}': "
                    f"{sheet_data.row_count} rows × {sheet_data.col_count} cols"
                )

            return ExcelDocument(
                file_path=str(file_path),
                file_name=file_path.name,
                file_size_bytes=file_size,
                sheet_names=list(sheets_data.keys()),
                sheets=sheets_data,
                named_ranges=named_ranges,
            )

        except Exception as exc:
            logger.error(f"Excel extraction failed: {exc}")
            return ExcelDocument(
                file_path=str(file_path),
                file_name=file_path.name,
                file_size_bytes=file_size,
                sheet_names=[],
                sheets={},
                named_ranges={},
                extraction_status="failed",
                error_message=str(exc),
            )

    def _extract_sheet(self, ws, sheet_name: str) -> SheetData:
        # Unmerge merged cells before reading
        has_merged = bool(ws.merged_cells.ranges)
        if has_merged:
            self._unmerge_cells(ws)

        # Read into DataFrame via pandas
        # Use openpyxl engine for consistent behavior
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))

        if not data:
            empty_df = pd.DataFrame()
            return SheetData(
                sheet_name=sheet_name,
                dataframe=empty_df,
                row_count=0,
                col_count=0,
                headers=[],
                has_merged_cells=has_merged,
            )

        # Detect header row
        header_row_idx = 0
        if self.header_row_detection:
            header_row_idx = self._detect_header_row(data)

        headers_raw = data[header_row_idx]
        headers = self._normalize_headers(headers_raw)
        rows = data[header_row_idx + 1:]

        # Adjust column count
        n_cols = len(headers)
        normalized_rows = [
            row[:n_cols] + [None] * max(0, n_cols - len(row))
            for row in rows
        ]

        df = pd.DataFrame(normalized_rows, columns=headers)
        df = df.dropna(how="all")  # Drop fully-empty rows

        if self.dtype_inference:
            df = self._infer_dtypes(df)

        return SheetData(
            sheet_name=sheet_name,
            dataframe=df,
            row_count=len(df),
            col_count=len(df.columns),
            headers=list(df.columns),
            has_merged_cells=has_merged,
        )

    def _detect_header_row(self, data: list[list], max_scan_rows: int = 10) -> int:
        """
        Heuristic: find the row where most cells are non-null strings
        and look like column names (no purely numeric values).
        """
        best_row = 0
        best_score = -1

        for i, row in enumerate(data[:max_scan_rows]):
            non_null = [c for c in row if c is not None]
            str_cells = [c for c in non_null if isinstance(c, str)]
            if not non_null:
                continue
            score = len(str_cells) / len(non_null)
            if score > best_score:
                best_score = score
                best_row = i

        return best_row

    def _normalize_headers(self, raw_headers: list) -> list[str]:
        """Clean header names: lowercase, snake_case, remove special chars."""
        headers = []
        seen: dict[str, int] = {}
        for i, h in enumerate(raw_headers):
            if h is None:
                name = f"col_{i}"
            else:
                name = re.sub(r"[^a-zA-Z0-9_ ]", "", str(h).strip())
                name = re.sub(r"\s+", "_", name).lower().strip("_")
                name = name or f"col_{i}"

            # Handle duplicates
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0

            headers.append(name)
        return headers

    def _infer_dtypes(self, df: pd.DataFrame) -> pd.DataFrame:
        """Attempt to cast object columns to numeric or datetime."""
        for col in df.columns:
            if df[col].dtype == object:
                # Try numeric
                try:
                    df[col] = pd.to_numeric(df[col])
                    continue
                except (ValueError, TypeError):
                    pass
                # Try datetime
                try:
                    df[col] = pd.to_datetime(df[col], infer_datetime_format=True)
                    continue
                except (ValueError, TypeError):
                    pass
        return df

    def _unmerge_cells(self, ws) -> None:
        """Fill merged cell ranges with the top-left value before reading."""
        for merge_range in list(ws.merged_cells.ranges):
            min_row, min_col = merge_range.min_row, merge_range.min_col
            top_left_value = ws.cell(min_row, min_col).value
            ws.unmerge_cells(str(merge_range))
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    ws.cell(row, col).value = top_left_value

    def _extract_named_ranges(self, wb) -> dict:
        named = {}
        for name, defn in wb.defined_names.items():
            try:
                destinations = list(defn.destinations)
                named[name] = [f"{sheet}!{cell}" for sheet, cell in destinations]
            except Exception:
                pass
        return named

    # ─── Convenience ─────────────────────────────────────────────────────────

    def extract_sheet_as_df(
        self, file_path: str | Path, sheet_name: str
    ) -> pd.DataFrame:
        """Quick helper to get a single sheet as DataFrame."""
        doc = self.extract(file_path)
        if sheet_name not in doc.sheets:
            raise KeyError(f"Sheet '{sheet_name}' not found in {file_path}")
        return doc.sheets[sheet_name].dataframe
